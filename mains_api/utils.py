from rest_framework.exceptions import ValidationError
from django.db import IntegrityError

from .models import *
import openpyxl


def save_clients(file):
    """Save Clients and Organization in DataBase"""
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.get_sheet_by_name('client')
    except (ValueError, KeyError):
        raise ValidationError(detail={'Error': 'Incorrect sheet or file name'})

    clients_objects = []
    for i in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[i] if cell.value is not None]

        if row:
            clients_objects.append(Clients(name=row[0]))

    try:
        Clients.objects.bulk_create(clients_objects)
    except IntegrityError:
        raise ValidationError(detail={'Error': 'Clients name already exists'})

    try:
        ws = wb.get_sheet_by_name('organization')
    except (ValueError, KeyError):
        raise ValidationError(detail={'Error': 'Incorrect sheet or file name'})

    if ws.max_row <= 1:
        raise ValidationError(detail={'Error': 'Empty filling'})

    organization_objects = []
    for i in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[i] if cell.value is not None]

        if row:
            organization_objects.append(Organization(
                client_name=row[0],
                name=row[1],
                address=row[2],
                client=Clients.objects.get(name=row[0]),
            ))

    try:
        Organization.objects.bulk_create(organization_objects)
    except IntegrityError:
        raise ValidationError(detail={'Error': 'Organization name already exists'})


def save_bills(file):
    """Save Bills in DataBase"""
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.get_sheet_by_name('Лист1')
    except (ValueError, KeyError):
        raise ValidationError(detail={'Error': 'Incorrect sheet or file name'})

    if ws.max_row <= 1:
        raise ValidationError(detail={'Error': 'Empty filling'})

    object_list = []
    for i in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[i] if cell.value is not None]

        if row:
            client_name = Clients.objects.get(name=row[0])
            client_org = Organization.objects.get(name=row[1])
            object_list.append(Bills(
                client_name=client_name,
                client_org=client_org,
                number=row[2],
                bills_sum=row[3],
                date=row[4],
                service=row[5]
            ))

    try:
        Bills.objects.bulk_create(object_list)
    except IntegrityError:
        raise ValidationError(detail={'Error': 'Client or organization name already exists'})
